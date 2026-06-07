# -*- coding: utf-8 -*-
"""
inspection_engine.py
موتور محاسباتی بازرسی پیش‌بسته‌بندی — ISIRI 16381 / OIML R87

ویژگی‌ها:
- مستقل از رابط کاربری؛ قابل استفاده در Streamlit / CLI / API
- کنترل واقعی الزامات پیش‌بسته‌بندی:
  1) رواداری TNE متناسب با مقدار اسمی
  2) درصد کمبود هر نمونه
  3) آزمون T1
  4) آزمون T2
  5) الزام میانگین تصحیح‌شده
  6) کنترل محاسبه وزن خالص از ناخالص و وزن بسته‌بندی
- تفکیک نتیجه اعتبارسنجی فایل از نتیجه استاندارد
- وابستگی خارجی فقط: openpyxl و scipy
"""

import math
import re
import statistics
import warnings
from pathlib import Path
from collections import Counter
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from scipy import stats


# ════════════════════════════ پیکربندی ════════════════════════════

class Config:
    """
    پارامترهای قابل تنظیم بازرسی.
    این کلاس را می‌توان از Streamlit / CLI / API مقداردهی کرد.
    """
    def __init__(
        self,
        min_rows: int = 20,
        near_dup_rel: float = 0.02,
        net_abs_tol: float = 0.01,
        percent_abs_tol: float = 0.05,
        tolerance_rel_tol: float = 0.01,
        strict_formula_coverage: bool = False,
    ):
        self.min_rows = min_rows
        self.near_dup_rel = near_dup_rel

        # تلورانس اختلاف وزن خالص محاسباتی با مقدار فرم
        self.net_abs_tol = net_abs_tol

        # تلورانس مقایسه درصد کمبود؛ واحد درصد
        self.percent_abs_tol = percent_abs_tol

        # تلورانس نسبی مقایسه رواداری فرم با رواداری استاندارد
        self.tolerance_rel_tol = tolerance_rel_tol

        # اگر True باشد، نبود فرمول در ستون‌های مهم ریسک بیشتری ایجاد می‌کند
        self.strict_formula_coverage = strict_formula_coverage


# ارقام فارسی و عربی به لاتین
PD = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")


COLUMN_KEYWORDS = {
    "gross": [
        "وزنناخالص", "ناخالص", "gross", "grossweight"
    ],
    "package": [
        "وزنبسته", "وزنبستهها", "وزنبستهبندی", "بستهها", "تار", "tare",
        "package", "packagingweight"
    ],
    "net_weight": [
        "وزنخالص", "خالصوزنی", "netweight", "netwt"
    ],
    "net_volume": [
        "حجمخالص", "خالصحجمی", "netvolume", "netvol"
    ],
    "t1_error": [
        "خطایt1", "آزمونt1", "t1", "t1error"
    ],
    "t2_error": [
        "خطایt2", "آزمونt2", "t2", "t2error"
    ],
    "shortage_pct": [
        "درصدکمبود", "درصدکسری", "درصدکسر", "کمبوددرصدی",
        "درصدکاهش", "درصدکمبودوزن", "درصدکمبودحجم",
        "shortagepercent", "shortagepct", "shortage"
    ],
    "tolerance": [
        "رواداری", "رواداریمجاز", "حدرواداری", "tolerance", "tne"
    ],
}


SUMMARY_PATTERNS = [
    ("nominal",         ["وزننامی", "حجمنامی", "مقدارنامی", "مقداراسمی", "مقداراظهارشد", "ظرفیتاسمی"]),
    ("tolerance",       ["رواداریمجاز", "رواداری", "حدرواداری", "tne"]),
    ("density",         ["چگالی", "دانسیته", "density"]),
    ("avg_net",         ["میانگینوزنخالص", "میانگینحجمخالص", "میانگینخالص"]),
    ("sd_net",          ["انحرافاستاندارد", "انحرافمعیار", "sd", "standarddeviation"]),
    ("corrected_avg",   ["میانگینوزنخالصتصحیح", "میانگینحجمخالصتصحیح", "میانگینتصحیح", "تصحیحشده"]),
    ("t1_fail_count",   ["تعدادبستههایردوددرآزمونt1", "تعدادردودt1", "مردودیt1"]),
    ("t2_fail_count",   ["تعدادبستههایردوددرآزمونt2", "تعدادردودt2", "مردودیt2"]),
    ("t1_result",       ["نتیجهآزمونt1"]),
    ("t2_result",       ["نتیجهآزمونt2"]),
    ("avg_requirement", ["آزمونالزاممیانگین", "الزاممیانگین"]),
    ("overall_result",  ["نتیجهکلی", "نتیجهبازرسی", "نتیجهنهایی"]),
]


RESULT_LABELS = {
    "file": "نام فایل",
    "sheet": "نام شیت",
    "n": "تعداد نمونه",

    "nominal": "مقدار اسمی",
    "mtype": "نوع اندازه‌گیری",
    "unit": "واحد",

    "file_verdict": "نتیجه اعتبارسنجی فایل",
    "standard_result": "نتیجه استاندارد",
    "form_result": "نتیجه فرم",
    "risk": "امتیاز ریسک",

    "tne": "رواداری استاندارد",
    "tne_pct": "رواداری درصدی استاندارد",
    "form_tolerance": "رواداری فرم",

    "mean": "میانگین",
    "sd": "SD",
    "corrected_avg": "میانگین تصحیح‌شده",
    "cv": "CV%",

    "t1_fail": "تعداد مردودی T1",
    "t1_allowed": "حد مجاز T1",
    "t2_fail": "تعداد مردودی T2",

    "avg_req": "نتیجه الزام میانگین",
    "t1_req": "نتیجه T1",
    "t2_req": "نتیجه T2",

    "net_calc_errors": "خطای محاسبه خالص",
    "shortage_errors": "خطای درصد کمبود",

    "dup_rate": "نرخ تکرار%",
    "blocks": "بلوک کپی",
    "iqr": "IQR پرت",
    "ex3": "۳σ پرت",
    "runs_p": "Runs p",
    "round_pct": "رقم رُند%",

    "standard_reasons": "دلایل استاندارد",
    "file_reasons": "دلایل اعتبارسنجی فایل",
}


# ════════════════════════════ جدول استاندارد پیش‌بسته‌بندی ════════════════════════════

def tne_table(nominal: float) -> Optional[float]:
    """
    حداکثر خطای مجاز منفی TNE بر اساس مقدار اسمی.
    واحد خروجی همان واحد مقدار اسمی است:
    گرم برای وزن، میلی‌لیتر برای حجم.
    """
    if nominal is None or nominal <= 0:
        return None

    n = nominal

    if n <= 50:
        return n * 0.09
    elif n <= 100:
        return 4.5
    elif n <= 200:
        return n * 0.045
    elif n <= 300:
        return 9.0
    elif n <= 500:
        return n * 0.03
    elif n <= 1000:
        return 15.0
    elif n <= 10000:
        return n * 0.015
    elif n <= 15000:
        return 150.0
    else:
        return n * 0.01


def correction_factor(n: int) -> Optional[float]:
    """
    ضریب تصحیح برای آزمون الزام میانگین.
    برای طرح‌های متداول:
    n≈50  -> 0.379
    n≈80  -> 0.295
    n≈125 -> 0.234

    برای n=20، مقدار 0.640 برای طرح تخریبی/کوچک در نظر گرفته شده است.
    """
    if n >= 120:
        return 0.234
    if n >= 80:
        return 0.295
    if n >= 50:
        return 0.379
    if n >= 20:
        return 0.640
    return None


def t1_allowed_count(n: int) -> Optional[int]:
    """
    تعداد مجاز بسته‌هایی که می‌توانند بین Qn-T و Qn-2T قرار بگیرند.
    بسته‌های کمتر از Qn-2T جداگانه با T2 کنترل می‌شوند.

    مقادیر متداول طرح نمونه‌برداری:
    n≈20  -> 1
    n≈50  -> 3
    n≈80  -> 5
    n≈125 -> 7
    """
    if n >= 120:
        return 7
    if n >= 80:
        return 5
    if n >= 50:
        return 3
    if n >= 20:
        return 1
    return None


# ════════════════════════════ توابع کمکی عمومی ════════════════════════════

def norm(s: Any) -> str:
    """نرمال‌سازی متن فارسی/انگلیسی برای جست‌وجوی کلیدواژه‌ها."""
    if s is None:
        return ""
    s = str(s).translate(PD)
    s = s.replace("ي", "ی").replace("ك", "ک").replace("\u200c", "")
    s = re.sub(r"[\s\-\–\—\(\)\[\]\{\}\.,:;،/\\]+", "", s)
    s = re.sub(r"[^\w\u0600-\u06FF]+", "", s)
    return s.lower()


def to_f(v: Any) -> Optional[float]:
    """
    تبدیل مقدار سلول به عدد.
    تلاش شده مشکل جداکننده هزارگان و اعشار کمتر شود.
    نمونه‌های قابل قبول:
    1000
    1,000
    1.5
    ۱٫۵
    ۱۲۳/۴۵ در صورت وجود ممیز فارسی/عربی تا حد ممکن
    """
    if v is None or v == "" or isinstance(v, bool):
        return None

    if isinstance(v, (int, float)):
        f = float(v)
        return None if math.isnan(f) else f

    s = str(v).translate(PD).strip()
    if not s:
        return None

    # ممیزهای عربی/فارسی احتمالی
    s = s.replace("٫", ".").replace("٬", ",")

    # پیدا کردن اولین توکن عددی
    m = re.search(r"[-+]?\d[\d,\.]*", s)
    if not m:
        return None

    token = m.group(0)

    # اگر هم کاما و هم نقطه وجود دارد، آخرین جداکننده را اعشار فرض می‌کنیم
    if "," in token and "." in token:
        last_comma = token.rfind(",")
        last_dot = token.rfind(".")
        dec_sep = "," if last_comma > last_dot else "."
        thou_sep = "." if dec_sep == "," else ","
        token = token.replace(thou_sep, "")
        token = token.replace(dec_sep, ".")
    elif "," in token:
        parts = token.split(",")
        if len(parts) > 2:
            token = "".join(parts[:-1]) + "." + parts[-1] if len(parts[-1]) != 3 else "".join(parts)
        else:
            # 1,000 را هزارگان، 12,5 را اعشار فرض می‌کنیم
            if len(parts[-1]) == 3 and parts[-1].isdigit():
                token = "".join(parts)
            else:
                token = token.replace(",", ".")
    elif "." in token:
        parts = token.split(".")
        if len(parts) > 2:
            # چند نقطه معمولاً هزارگان است
            token = "".join(parts)
        else:
            token = token

    try:
        return float(token)
    except Exception:
        return None


def fmt(x: Any, nd: int = 4) -> str:
    if x is None:
        return "—"
    try:
        x = float(x)
    except Exception:
        return str(x)
    if math.isnan(x):
        return "—"
    s = f"{x:.{nd}f}".rstrip("0").rstrip(".")
    return s or "0"


def sdev(xs: List[float]) -> Optional[float]:
    return statistics.stdev(xs) if len(xs) >= 2 else None


def mean(xs: List[float]) -> Optional[float]:
    return statistics.fmean(xs) if xs else None


def has_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def percentile(xs: List[float], p: float) -> Optional[float]:
    """صدک با درون‌یابی خطی؛ جایگزین ساده برای numpy.percentile."""
    if not xs:
        return None
    ys = sorted(xs)
    if len(ys) == 1:
        return ys[0]
    k = (len(ys) - 1) * (p / 100)
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return ys[int(k)]
    return ys[f] + (ys[c] - ys[f]) * (k - f)


def boolish(v: Any) -> Optional[bool]:
    """
    تبدیل مقدار سلول به True/False در صورت امکان.
    برای کنترل ستون‌های T1/T2 که ممکن است متنی یا عددی باشند.
    """
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v

    f = to_f(v)
    if f is not None:
        if abs(f) < 1e-12:
            return False
        if abs(f - 1) < 1e-12:
            return True

    t = norm(v)

    true_words = [
        "true", "yes", "y", "بله", "دارد", "مردود", "رد", "نامنطبق",
        "خطادارد", "fail", "failed", "reject", "rejected"
    ]
    false_words = [
        "false", "no", "n", "خیر", "ندارد", "قبول", "تایید", "منطبق",
        "pass", "passed", "accept", "accepted", "ok"
    ]

    if any(w in t for w in true_words):
        return True
    if any(w in t for w in false_words):
        return False

    return None


def close_enough(a: float, b: float, abs_tol: float = 1e-9, rel_tol: float = 1e-6) -> bool:
    return abs(a - b) <= max(abs_tol, abs(b) * rel_tol)


# ════════════════════════════ کشف ساختار فایل اکسل ════════════════════════════

def find_header(ws, scan: int = 25) -> int:
    kws = [
        "وزنناخالص", "وزنبسته", "وزنبستهبندی", "وزنخالص", "حجمخالص",
        "خطایt1", "خطایt2", "نامفراورده", "درصدکمبود", "رواداری"
    ]

    best_score = -1
    best_row = 1

    for r in range(1, min(ws.max_row, scan) + 1):
        row_text = " ".join(
            norm(ws.cell(r, c).value)
            for c in range(1, min(ws.max_column, 25) + 1)
        )
        score = sum(1 for k in kws if k in row_text)

        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def col_map(ws, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}

    for c in range(1, min(ws.max_column, 35) + 1):
        t = norm(ws.cell(header_row, c).value)
        if not t:
            continue

        for key, kws in COLUMN_KEYWORDS.items():
            if key not in m and any(k in t for k in kws):
                m[key] = c

    return m


def sample_end(ws, header_row: int, cm: Dict[str, int]) -> int:
    cols = [
        cm[k] for k in (
            "gross", "package", "net_weight", "net_volume",
            "t1_error", "t2_error", "shortage_pct", "tolerance"
        )
        if k in cm
    ]

    if not cols:
        cols = list(range(1, min(ws.max_column, 10) + 1))

    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in cols):
            last = r

    return last


def series(wsv, c: int, s: int, e: int) -> List[float]:
    xs = []
    for r in range(s, e + 1):
        v = to_f(wsv.cell(r, c).value)
        if v is not None:
            xs.append(v)
    return xs


def series_rows(wsv, c: int, s: int, e: int) -> List[Tuple[int, float]]:
    xs = []
    for r in range(s, e + 1):
        v = to_f(wsv.cell(r, c).value)
        if v is not None:
            xs.append((r, v))
    return xs


def adjacent_value(wsv, r: int, c: int) -> Any:
    """
    پیدا کردن مقدار مربوط به یک عنوان خلاصه.
    در فرم‌های راست‌به‌چپ معمولاً مقدار در خانه سمت چپ عنوان است.
    اما برای انعطاف، چند خانه اطراف بررسی می‌شود.
    """
    candidates = [
        (r, c - 1),
        (r, c + 1),
        (r + 1, c),
        (r + 1, c - 1),
        (r + 1, c + 1),
    ]

    for rr, cc in candidates:
        if rr < 1 or cc < 1:
            continue
        v = wsv.cell(rr, cc).value
        if v not in (None, ""):
            return v

    return None


def find_summary(ws, wsv) -> Dict[str, Dict[str, Any]]:
    found: Dict[str, Dict[str, Any]] = {}

    for r in range(1, min(ws.max_row, 50) + 1):
        for c in range(1, min(ws.max_column, 18) + 1):
            kt = norm(ws.cell(r, c).value)
            if not kt:
                continue

            for key, kws in SUMMARY_PATTERNS:
                if key not in found and any(k in kt for k in kws):
                    found[key] = {
                        "value": adjacent_value(wsv, r, c),
                        "row": r,
                        "col": c,
                        "label": ws.cell(r, c).value,
                    }

    return found


# ════════════════════════════ آزمون‌های آماری کمکی برای اعتبارسنجی فایل ════════════════════════════

def runs_test(xs: List[float]) -> Optional[Dict[str, float]]:
    if len(xs) < 12:
        return None

    med = statistics.median(xs)
    signs = [1 if x > med else -1 for x in xs if x != med]

    n1 = signs.count(1)
    n2 = signs.count(-1)

    if n1 == 0 or n2 == 0:
        return None

    runs = 1 + sum(1 for i in range(1, len(signs)) if signs[i] != signs[i - 1])
    mu = 1 + 2 * n1 * n2 / (n1 + n2)
    var = (
        2 * n1 * n2 * (2 * n1 * n2 - n1 - n2)
        / ((n1 + n2) ** 2 * (n1 + n2 - 1))
    )

    if var <= 0:
        return None

    z = (runs - mu) / math.sqrt(var)
    return {"p": 2 * (1 - stats.norm.cdf(abs(z)))}


def duplicate_stats(xs: List[float]) -> Dict[str, float]:
    c = Counter(xs)
    dup = sum(v - 1 for v in c.values() if v > 1)
    return {
        "dup_rate": dup / len(xs) if xs else 0,
        "max_repeat": max(c.values()) if c else 0,
    }


def repeated_blocks(xs: List[float], blk: int = 3) -> int:
    if len(xs) < blk * 2:
        return 0

    seen = {}
    hits = 0

    for i in range(len(xs) - blk + 1):
        key = tuple(round(v, 4) for v in xs[i:i + blk])
        if key in seen:
            hits += 1
        else:
            seen[key] = i

    return hits


def decimal_round_bias(xs: List[float]) -> Optional[Dict[str, float]]:
    if not xs:
        return None

    last = []
    for x in xs:
        s = f"{x:.1f}"
        if "." in s:
            last.append(s.split(".")[1][-1])
        else:
            last.append("0")

    c = Counter(last)

    return {
        "pct_0_5": (c.get("0", 0) + c.get("5", 0)) / len(last) * 100
    }


def iqr_outliers(xs: List[float]) -> Dict[str, int]:
    if len(xs) < 8:
        return {"count": 0}

    q1 = percentile(xs, 25)
    q3 = percentile(xs, 75)

    if q1 is None or q3 is None:
        return {"count": 0}

    iqr = q3 - q1
    lo = q1 - 1.5 * iqr
    hi = q3 + 1.5 * iqr

    return {
        "count": sum(1 for x in xs if x < lo or x > hi)
    }


def extreme_3sigma(xs: List[float]) -> int:
    if len(xs) < 2:
        return 0

    m = mean(xs)
    sd = sdev(xs)

    if m is None or not sd:
        return 0

    return sum(1 for x in xs if abs(x - m) > 3 * sd)


def too_perfect_normal(xs: List[float]) -> Optional[Dict[str, float]]:
    if len(xs) < 30 or len(xs) > 5000:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        w, p = stats.shapiro(xs)

    return {"W": float(w), "p": float(p)}


def grid_regularity(xs: List[float]) -> float:
    if len(xs) < 10:
        return 0.0

    uniq = sorted(set(xs))
    if len(uniq) < 4:
        return 0.0

    diffs = [
        round(uniq[i] - uniq[i - 1], 4)
        for i in range(1, len(uniq))
    ]
    diffs = [d for d in diffs if d > 0]

    if len(diffs) < 3:
        return 0.0

    c = Counter(diffs)
    return c.most_common(1)[0][1] / len(diffs) * 100


# ════════════════════════════ تشخیص نوع اندازه‌گیری ════════════════════════════

def detect_type(cm: Dict[str, int], summary: Dict[str, Dict[str, Any]]) -> str:
    has_vol = "net_volume" in cm
    has_wt = "net_weight" in cm

    if has_vol and not has_wt:
        return "حجم"

    if has_wt and not has_vol:
        return "وزن"

    if "density" in summary:
        return "حجم"

    return "حجم" if has_vol else "وزن"


# ════════════════════════════ محاسبات واقعی استاندارد ════════════════════════════

def independent_standard_check(
    net_vals: List[float],
    nominal: Optional[float],
    n: int
) -> Dict[str, Any]:
    """
    کنترل مستقل استاندارد پیش‌بسته‌بندی.
    خروجی این تابع نباید به نتایج نوشته‌شده در فرم اکسل وابسته باشد.
    """
    out: Dict[str, Any] = {
        "tne": None,
        "tne_pct": None,
        "mean": None,
        "sd": None,
        "corrected_avg": None,
        "t1_fail": None,
        "t2_fail": None,
        "t1_allowed": None,
        "avg_req": "نامشخص",
        "t1_req": "نامشخص",
        "t2_req": "نامشخص",
        "standard_result": "نامشخص",
        "standard_reasons": [],
    }

    if nominal is None or nominal <= 0:
        out["standard_reasons"].append("مقدار اسمی معتبر یافت نشد.")
        return out

    if not net_vals:
        out["standard_reasons"].append("داده خالص برای آزمون استاندارد یافت نشد.")
        return out

    tne = tne_table(nominal)
    out["tne"] = tne
    out["tne_pct"] = (tne / nominal * 100) if tne else None

    m = mean(net_vals)
    sd = sdev(net_vals)
    k = correction_factor(n)
    allowed = t1_allowed_count(n)

    out["mean"] = m
    out["sd"] = sd
    out["t1_allowed"] = allowed

    if sd is not None and k is not None and m is not None:
        out["corrected_avg"] = m - k * sd
        out["avg_req"] = "قبول" if out["corrected_avg"] >= nominal else "مردود"
    else:
        out["avg_req"] = "نامشخص"
        out["standard_reasons"].append("به دلیل تعداد نمونه/انحراف معیار، آزمون میانگین کامل نیست.")

    if tne is not None:
        out["t1_fail"] = sum(1 for x in net_vals if x < nominal - tne)
        out["t2_fail"] = sum(1 for x in net_vals if x < nominal - 2 * tne)

        if allowed is not None:
            out["t1_req"] = "قبول" if out["t1_fail"] <= allowed else "مردود"
        else:
            out["t1_req"] = "نامشخص"
            out["standard_reasons"].append("تعداد نمونه برای تعیین حد مجاز T1 کافی نیست.")

        out["t2_req"] = "قبول" if out["t2_fail"] == 0 else "مردود"

    if out["avg_req"] == "مردود":
        out["standard_reasons"].append(
            f"الزام میانگین مردود است: میانگین تصحیح‌شده={fmt(out['corrected_avg'])} < مقدار اسمی={fmt(nominal)}"
        )

    if out["t1_req"] == "مردود":
        out["standard_reasons"].append(
            f"تعداد مردودی T1 بیش از حد مجاز است: {out['t1_fail']} > {out['t1_allowed']}"
        )

    if out["t2_req"] == "مردود":
        out["standard_reasons"].append(
            f"وجود بسته کمتر از Qn-2T مجاز نیست: تعداد T2={out['t2_fail']}"
        )

    decisive = [out["avg_req"], out["t1_req"], out["t2_req"]]

    if any(x == "مردود" for x in decisive):
        out["standard_result"] = "مردود"
    elif all(x == "قبول" for x in decisive):
        out["standard_result"] = "قبول"
    else:
        out["standard_result"] = "نامشخص"

    if not out["standard_reasons"] and out["standard_result"] == "قبول":
        out["standard_reasons"].append("هر سه شرط اصلی استاندارد برقرار است.")

    return out


def shortage_percent(nominal: float, actual: float) -> float:
    """درصد کمبود نسبت به مقدار اسمی. اضافه‌بودن، کمبود محسوب نمی‌شود."""
    if nominal <= 0:
        return 0.0
    return max(0.0, (nominal - actual) / nominal * 100)


# ════════════════════════════ پردازش یک شیت ════════════════════════════

def process_sheet(ws, wsv, fname: str, sname: str, cfg: Config) -> Dict[str, Any]:
    hr = find_header(ws)
    cm = col_map(ws, hr)
    se = sample_end(ws, hr, cm)
    ss = hr + 1
    n = max(0, se - ss + 1)

    summary = find_summary(ws, wsv)

    mtype = detect_type(cm, summary)
    net_key = "net_weight" if mtype == "وزن" else "net_volume"
    unit = "گرم" if mtype == "وزن" else "میلی‌لیتر"

    file_reasons: List[str] = []
    standard_extra_reasons: List[str] = []

    risk = 0
    critical = False

    # کنترل وجود ستون‌های پایه
    if "gross" not in cm and mtype == "وزن":
        file_reasons.append("ستون وزن ناخالص یافت نشد.")
        risk += 15

    if net_key not in cm:
        critical = True
        file_reasons.append(f"ستون {mtype} خالص یافت نشد.")
        risk += 30

    for k, lbl in [("t1_error", "خطای T1"), ("t2_error", "خطای T2")]:
        if k not in cm:
            file_reasons.append(f"ستون {lbl} یافت نشد.")
            risk += 5

    # مقدار اسمی
    nominal = to_f(summary.get("nominal", {}).get("value"))

    if nominal is None:
        m = re.search(r"(\d+(?:[\.,]\d+)?)", str(sname).translate(PD))
        if m:
            nominal = to_f(m.group(1))

    # سری اصلی
    primary = series(wsv, cm[net_key], ss, se) if net_key in cm else []
    primary_rows = series_rows(wsv, cm[net_key], ss, se) if net_key in cm else []

    sp = {
        "mean": mean(primary),
        "sd": sdev(primary),
        "cv": None,
    }

    if sp["mean"] and sp["sd"]:
        sp["cv"] = sp["sd"] / sp["mean"] * 100

    # کنترل مستقل استاندارد
    iso = independent_standard_check(primary, nominal, n)

    # رواداری فرم در خلاصه
    form_tol = to_f(summary.get("tolerance", {}).get("value"))
    tne = iso.get("tne")
    tne_pct = iso.get("tne_pct")

    if nominal and tne is not None:
        if form_tol is not None:
            # حالت اصلی: رواداری فرم باید با مقدار مطلق TNE برابر باشد.
            if not close_enough(form_tol, tne, abs_tol=cfg.net_abs_tol, rel_tol=cfg.tolerance_rel_tol):
                # اگر فرم به جای مقدار مطلق، درصد رواداری را نوشته باشد، گزارش جداگانه بده
                if tne_pct is not None and close_enough(form_tol, tne_pct, abs_tol=cfg.percent_abs_tol, rel_tol=cfg.tolerance_rel_tol):
                    standard_extra_reasons.append(
                        f"رواداری فرم ظاهراً به درصد نوشته شده است: فرم={fmt(form_tol)}٪، مقدار مطلق استاندارد={fmt(tne)} {unit}"
                    )
                    risk += 5
                else:
                    standard_extra_reasons.append(
                        f"رواداری فرم با جدول استاندارد تطابق ندارد: فرم={fmt(form_tol)}، استاندارد={fmt(tne)} {unit}"
                    )
                    risk += 25
        else:
            file_reasons.append("رواداری مجاز در بخش خلاصه فرم یافت نشد.")
            risk += 8

    # مقایسه خلاصه‌های فرم با محاسبات مستقل
    def cross_check(skey: str, calc: Optional[float], label: str, tol_rel: float = 0.02, tol_abs: float = 0.01):
        nonlocal risk

        rec = to_f(summary.get(skey, {}).get("value"))

        if rec is None or calc is None:
            return

        if abs(rec - calc) > max(tol_abs, abs(calc) * tol_rel):
            file_reasons.append(f"{label}: فرم={fmt(rec)} ≠ محاسبه={fmt(calc)}")
            risk += 20 if abs(rec - calc) > max(tol_abs, abs(calc) * 0.05) else 10

    cross_check("avg_net", iso.get("mean"), f"میانگین {mtype} خالص")
    cross_check("sd_net", iso.get("sd"), "انحراف معیار")
    cross_check("corrected_avg", iso.get("corrected_avg"), "میانگین تصحیح‌شده")
    cross_check("t1_fail_count", iso.get("t1_fail"), "تعداد مردودی T1", tol_rel=0, tol_abs=0)
    cross_check("t2_fail_count", iso.get("t2_fail"), "تعداد مردودی T2", tol_rel=0, tol_abs=0)

    # کنترل محاسبه وزن خالص = ناخالص - وزن بسته‌بندی
    net_calc_errors = 0

    if mtype == "وزن" and all(k in cm for k in ("gross", "package", "net_weight")):
        for r in range(ss, se + 1):
            gross = to_f(wsv.cell(r, cm["gross"]).value)
            tare = to_f(wsv.cell(r, cm["package"]).value)
            net = to_f(wsv.cell(r, cm["net_weight"]).value)

            if gross is None or tare is None or net is None:
                continue

            calc_net = gross - tare

            if abs(calc_net - net) > cfg.net_abs_tol:
                net_calc_errors += 1

        if net_calc_errors:
            file_reasons.append(
                f"{net_calc_errors} ردیف اختلاف در محاسبه وزن خالص = ناخالص - وزن بسته‌بندی"
            )
            risk += min(30, net_calc_errors * 3)

    elif mtype == "وزن":
        file_reasons.append("برای کنترل وزن خالص، ستون ناخالص یا وزن بسته‌بندی کامل نیست.")
        risk += 5

    # کنترل درصد کمبود و T1/T2 ردیفی
    shortage_errors = 0
    t1_col_mismatch = 0
    t2_col_mismatch = 0
    tolerance_col_mismatch = 0

    if nominal and tne is not None and primary_rows:
        for r, actual in primary_rows:
            calc_short = shortage_percent(nominal, actual)
            calc_t1 = actual < nominal - tne
            calc_t2 = actual < nominal - 2 * tne

            # ستون درصد کمبود
            if "shortage_pct" in cm:
                rec_short = to_f(wsv.cell(r, cm["shortage_pct"]).value)
                if rec_short is not None:
                    if abs(rec_short - calc_short) > cfg.percent_abs_tol:
                        shortage_errors += 1

            # ستون رواداری ردیفی، اگر وجود داشته باشد
            if "tolerance" in cm:
                rec_tol = to_f(wsv.cell(r, cm["tolerance"]).value)
                if rec_tol is not None:
                    if not close_enough(rec_tol, tne, abs_tol=cfg.net_abs_tol, rel_tol=cfg.tolerance_rel_tol):
                        # احتمال اینکه درصد باشد
                        if tne_pct is None or not close_enough(rec_tol, tne_pct, abs_tol=cfg.percent_abs_tol, rel_tol=cfg.tolerance_rel_tol):
                            tolerance_col_mismatch += 1

            # ستون T1
            if "t1_error" in cm:
                rec_t1 = boolish(wsv.cell(r, cm["t1_error"]).value)
                if rec_t1 is not None and rec_t1 != calc_t1:
                    t1_col_mismatch += 1

            # ستون T2
            if "t2_error" in cm:
                rec_t2 = boolish(wsv.cell(r, cm["t2_error"]).value)
                if rec_t2 is not None and rec_t2 != calc_t2:
                    t2_col_mismatch += 1

    if "shortage_pct" not in cm:
        file_reasons.append("ستون درصد کمبود وزن/حجم یافت نشد.")
        risk += 5
    elif shortage_errors:
        file_reasons.append(f"{shortage_errors} ردیف اختلاف در درصد کمبود درج‌شده با محاسبه مستقل")
        risk += min(25, shortage_errors * 3)

    if "tolerance" in cm and tolerance_col_mismatch:
        standard_extra_reasons.append(
            f"{tolerance_col_mismatch} ردیف دارای رواداری نامتناسب با مقدار اسمی است."
        )
        risk += min(25, tolerance_col_mismatch * 3)

    if t1_col_mismatch:
        file_reasons.append(f"{t1_col_mismatch} ردیف اختلاف بین ستون T1 فرم و محاسبه مستقل")
        risk += min(20, t1_col_mismatch * 2)

    if t2_col_mismatch:
        file_reasons.append(f"{t2_col_mismatch} ردیف اختلاف بین ستون T2 فرم و محاسبه مستقل")
        risk += min(25, t2_col_mismatch * 3)

    # کنترل حجم/وزن با چگالی اگر موجود باشد
    density = to_f(summary.get("density", {}).get("value"))

    if mtype == "حجم" and density and "net_weight" in cm and "net_volume" in cm:
        wt = series(wsv, cm["net_weight"], ss, se)
        vol = series(wsv, cm["net_volume"], ss, se)

        if wt and vol and len(wt) == len(vol):
            diffs = []
            for i in range(len(wt)):
                if vol[i]:
                    diffs.append(abs(vol[i] - wt[i] / density) / abs(vol[i]))

            if diffs and mean(diffs) and mean(diffs) > 0.01:
                file_reasons.append(
                    f"ناهماهنگی حجم↔وزن/چگالی، میانگین اختلاف={fmt(mean(diffs) * 100, 1)}٪"
                )
                risk += 15

    # کنترل پوشش فرمول‌ها
    for k, lbl in [(net_key, "مقدار خالص"), ("t1_error", "T1"), ("t2_error", "T2")]:
        if k not in cm:
            continue

        c = cm[k]
        fr = sum(
            1 for r in range(ss, se + 1)
            if to_f(wsv.cell(r, c).value) is not None and has_formula(ws.cell(r, c))
        )

        cov = fr / n if n else 0

        if cfg.strict_formula_coverage and cov < 0.9:
            file_reasons.append(f"پوشش فرمول {lbl}: {cov * 100:.0f}%")
            risk += int((1 - cov) * (30 if k == net_key else 15))
        elif cov < 0.5 and k == net_key:
            file_reasons.append(f"پوشش فرمول {lbl} پایین است: {cov * 100:.0f}%")
            risk += 8

    # آزمون‌های آماری/اعتبارسنجی داده
    dup = None
    run = None
    dec = None
    iqr_c = 0
    ex3 = 0
    blocks = 0
    grid = 0.0

    if primary:
        dup = duplicate_stats(primary)
        run = runs_test(primary)
        dec = decimal_round_bias(primary)
        iqr_c = iqr_outliers(primary)["count"]
        ex3 = extreme_3sigma(primary)
        blocks = repeated_blocks(primary)
        grid = grid_regularity(primary)
        norm_t = too_perfect_normal(primary)

        if dup["dup_rate"] > 0.4:
            file_reasons.append(f"نرخ تکرار بالا: {dup['dup_rate'] * 100:.0f}%")
            risk += 8

        if dup["max_repeat"] >= max(5, n * 0.1):
            file_reasons.append(f"یک مقدار {dup['max_repeat']} بار تکرار شده")
            risk += 6

        if blocks > n * 0.15:
            file_reasons.append(f"{blocks} بلوک متوالی کپی‌شده")
            risk += 12

        if grid and grid > 70:
            file_reasons.append(f"شبکه تفاضلی بیش‌ازحد منظم: {grid:.0f}%")
            risk += 10

        if run and run["p"] < 0.01:
            file_reasons.append(f"الگوی غیرتصادفی، Runs p={run['p']:.3g}")
            risk += 5

        if dec and dec["pct_0_5"] > 55:
            file_reasons.append(f"تمایل به رقم رُند ۰/۵: {dec['pct_0_5']:.0f}%")
            risk += 6

        if iqr_c > 0:
            file_reasons.append(f"{iqr_c} داده پرت بر اساس IQR")
            risk += min(10, iqr_c * 3)

        if ex3 > 0:
            file_reasons.append(f"{ex3} داده فراتر از ۳σ؛ احتمال خطای تایپ")
            risk += min(20, ex3 * 8)

        if norm_t and norm_t["p"] > 0.95 and dup["dup_rate"] < 0.05:
            file_reasons.append("نرمال‌بودن بیش‌ازحد عالی؛ احتمال ساخت مصنوعی داده")
            risk += 8

    # کنترل فاصله شدید میانگین از اسمی
    if nominal and sp["mean"] and abs(sp["mean"] - nominal) / nominal > 0.1:
        file_reasons.append("میانگین بیش از ۱۰٪ از مقدار اسمی فاصله دارد.")
        risk += 15

    # نتیجه کلی فرم
    ov = summary.get("overall_result", {}).get("value")
    overall = None

    if ov is not None:
        b = boolish(ov)
        if b is not None:
            overall = b
        else:
            t = norm(ov)
            if "قبول" in t or "تایید" in t or "منطبق" in t:
                overall = True
            elif "مردود" in t or "رد" in t or "نامنطبق" in t:
                overall = False

    if overall is False:
        file_reasons.append("نتیجه کلی فرم: مردود")
        risk += 10

    # مقایسه نتیجه فرم با نتیجه استاندارد مستقل
    if overall is not None and iso.get("standard_result") in ("قبول", "مردود"):
        form_res_text = "قبول" if overall else "مردود"
        if form_res_text != iso["standard_result"]:
            standard_extra_reasons.append(
                f"نتیجه فرم با محاسبه مستقل استاندارد همخوان نیست: فرم={form_res_text}، محاسبه={iso['standard_result']}"
            )
            risk += 25

    # کنترل حداقل تعداد نمونه
    if n < cfg.min_rows:
        critical = True
        file_reasons.append(f"تعداد نمونه کافی نیست: {n}")
        risk += 20

    if not primary:
        critical = True
        file_reasons.append("ستون عددی ملاک یافت نشد یا داده عددی ندارد.")
        risk += 30

    # دلایل استاندارد نهایی
    standard_reasons = list(iso.get("standard_reasons", []))
    standard_reasons.extend(standard_extra_reasons)

    if not standard_reasons:
        standard_reasons.append("مورد استانداردی قابل گزارش مشاهده نشد.")

    risk = min(100, int(round(risk)))

    file_verdict = (
        "خطا" if critical
        else "بسیار مشکوک" if risk >= 65
        else "مشکوک" if risk >= 35
        else "سالم"
    )

    L = RESULT_LABELS

    return {
        L["file"]: fname,
        L["sheet"]: sname,
        L["n"]: n,

        L["nominal"]: fmt(nominal) if nominal else "نامشخص",
        L["mtype"]: mtype,
        L["unit"]: unit,

        L["file_verdict"]: file_verdict,
        L["standard_result"]: iso.get("standard_result", "نامشخص"),
        L["form_result"]: "قبول" if overall else "مردود" if overall is False else "نامشخص",
        L["risk"]: risk,

        L["tne"]: fmt(iso.get("tne")),
        L["tne_pct"]: fmt(iso.get("tne_pct"), 3),
        L["form_tolerance"]: fmt(form_tol),

        L["mean"]: fmt(iso.get("mean")),
        L["sd"]: fmt(iso.get("sd")),
        L["corrected_avg"]: fmt(iso.get("corrected_avg")),
        L["cv"]: fmt(sp["cv"], 2),

        L["t1_fail"]: iso.get("t1_fail") if iso.get("t1_fail") is not None else "—",
        L["t1_allowed"]: iso.get("t1_allowed") if iso.get("t1_allowed") is not None else "—",
        L["t2_fail"]: iso.get("t2_fail") if iso.get("t2_fail") is not None else "—",

        L["avg_req"]: iso.get("avg_req", "نامشخص"),
        L["t1_req"]: iso.get("t1_req", "نامشخص"),
        L["t2_req"]: iso.get("t2_req", "نامشخص"),

        L["net_calc_errors"]: net_calc_errors,
        L["shortage_errors"]: shortage_errors,

        L["dup_rate"]: fmt(dup["dup_rate"] * 100, 1) if dup else "—",
        L["blocks"]: blocks,
        L["iqr"]: iqr_c,
        L["ex3"]: ex3,
        L["runs_p"]: fmt(run["p"], 4) if run else "—",
        L["round_pct"]: fmt(dec["pct_0_5"], 1) if dec else "—",

        L["standard_reasons"]: " | ".join(standard_reasons),
        L["file_reasons"]: " | ".join(file_reasons) if file_reasons else "موردی مشاهده نشد.",
    }


# ════════════════════════════ API عمومی موتور ════════════════════════════

def process_workbook(path, cfg: Config) -> List[Dict[str, Any]]:
    """
    پردازش یک فایل اکسل.
    خروجی: لیست رکوردهای همه شیت‌ها.
    """
    path = Path(path)

    wb = load_workbook(path, data_only=False)
    wv = load_workbook(path, data_only=True)

    records: List[Dict[str, Any]] = []

    for s in wb.sheetnames:
        records.append(
            process_sheet(wb[s], wv[s], path.name, s, cfg)
        )

    return records


def error_record(fname: str, msg: str) -> Dict[str, Any]:
    """
    رکورد خطا برای فایل‌هایی که باز یا پردازش نمی‌شوند.
    """
    L = RESULT_LABELS
    base = {v: "—" for v in RESULT_LABELS.values()}

    base.update({
        L["file"]: fname,
        L["sheet"]: "خطا",
        L["n"]: 0,
        L["file_verdict"]: "خطا",
        L["standard_result"]: "نامشخص",
        L["form_result"]: "نامشخص",
        L["risk"]: 100,
        L["blocks"]: 0,
        L["iqr"]: 0,
        L["ex3"]: 0,
        L["standard_reasons"]: "به دلیل خطای پردازش، آزمون استاندارد انجام نشد.",
        L["file_reasons"]: f"خطا در پردازش: {msg}",
    })

    return base


def build_report(records: List[Dict[str, Any]], out_path) -> Path:
    """
    ساخت گزارش Excel رنگی.
    """
    out_path = Path(out_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش"

    headers = list(records[0].keys()) if records else []
    ws.append(headers)

    for r in records:
        ws.append([r.get(h) for h in headers])

    # هدر
    hf = PatternFill("solid", fgColor="0B1E3D")
    for c in ws[1]:
        c.fill = hf
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # رنگ‌بندی بر اساس نتیجه استاندارد، اگر موجود باشد
    std_header = RESULT_LABELS["standard_result"]
    file_header = RESULT_LABELS["file_verdict"]

    std_idx = headers.index(std_header) if std_header in headers else None
    file_idx = headers.index(file_header) if file_header in headers else None

    std_fills = {
        "قبول": "D1FAE5",
        "مردود": "FEE2E2",
        "نامشخص": "E5E7EB",
    }

    file_fills = {
        "سالم": "D1FAE5",
        "مشکوک": "FEF3C7",
        "بسیار مشکوک": "FEE2E2",
        "خطا": "FEE2E2",
    }

    for row in ws.iter_rows(min_row=2):
        fill_color = None

        if std_idx is not None:
            fill_color = std_fills.get(row[std_idx].value)

        if fill_color is None and file_idx is not None:
            fill_color = file_fills.get(row[file_idx].value)

        if fill_color:
            for c in row:
                c.fill = PatternFill("solid", fgColor=fill_color)

    for i, h in enumerate(headers, 1):
        if h in (RESULT_LABELS["standard_reasons"], RESULT_LABELS["file_reasons"]):
            width = 70
        elif h in (RESULT_LABELS["file"], RESULT_LABELS["sheet"]):
            width = 25
        else:
            width = 16

        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = True

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(out_path)
    return out_path


# ════════════════════════════ CLI اختیاری ════════════════════════════

def _cli():
    import argparse
    import glob

    ap = argparse.ArgumentParser(
        description="موتور بازرسی پیش‌بسته‌بندی ISIRI 16381 / OIML R87"
    )

    ap.add_argument(
        "inputs",
        nargs="+",
        help="فایل‌های ورودی .xlsx یا الگو، مثل *.xlsx"
    )

    ap.add_argument(
        "-o",
        "--output",
        default="report.xlsx",
        help="مسیر گزارش خروجی"
    )

    ap.add_argument(
        "--min-rows",
        type=int,
        default=20,
        help="حداقل تعداد نمونه"
    )

    args = ap.parse_args()

    cfg = Config(min_rows=args.min_rows)

    paths: List[str] = []
    for pat in args.inputs:
        paths.extend(glob.glob(pat))

    records: List[Dict[str, Any]] = []

    for p in paths:
        try:
            records.extend(process_workbook(p, cfg))
            print(f"✔ پردازش شد: {p}")
        except Exception as e:
            records.append(error_record(Path(p).name, str(e)))
            print(f"✖ خطا در {p}: {e}")

    if records:
        build_report(records, args.output)
        print(f"\n📄 گزارش ذخیره شد: {args.output}  ({len(records)} شیت)")
    else:
        print("هیچ فایلی پردازش نشد.")


if __name__ == "__main__":
    _cli()