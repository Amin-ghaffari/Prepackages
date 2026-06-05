# -*- coding: utf-8 -*-
"""
بازرسی پیشرفته فایل‌های اکسل - ISIRI 16381
نسخه با UI حرفه‌ای
"""

import streamlit as st
import pandas as pd
from pathlib import Path
import tempfile
from openpyxl import Workbook
import math, re, statistics, warnings
from collections import Counter
from typing import Any, Dict, List, Optional, Tuple
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from scipy import stats

# ──────────────────────────────────────────────
#  تنظیمات اصلی
# ──────────────────────────────────────────────
DEFAULT_NOMINAL = None
MIN_ROWS_FOR_ANALYSIS = 20

PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")

COLUMN_KEYWORDS = {
    "gross":             ["وزنناخالص", "gross"],
    "package":           ["وزنبسته", "بستهها", "tare", "package"],
    "net_weight":        ["وزنخالص", "netweight"],
    "net_volume":        ["حجمخالص", "netvolume"],
    "corrected_weight":  ["وزنخالصتصحیح", "correctedweight", "وزن تصحیح"],
    "corrected_volume":  ["حجمخالصتصحیح", "correctedvolume", "حجم تصحیح"],
    "t1_error":          ["خطایt1", "error1", "t1"],
    "t2_error":          ["خطایt2", "error2", "t2"],
}

SUMMARY_PATTERNS = [
    ("nominal",          ["وزننامی", "حجمنامی", "مقدارنامی"]),
    ("tolerance",        ["رواداریمجاز"]),
    ("density",          ["چگالی"]),
    ("avg_package",      ["میانگینوزنبسته"]),
    ("avg_net",          ["میانگینوزنخالص", "میانگینحجمخالص"]),
    ("sd_net",           ["انحرافاستانداردوزن", "انحرافاستانداردحجم", "انحرافمعیار"]),
    ("corrected_avg",    ["میانگینوزنخالصتصحیح", "میانگینحجمخالصتصحیح"]),
    ("t1_fail_count",    ["تعدادبستههایردوددرآزمونt1"]),
    ("t2_fail_count",    ["تعدادبستههایردوددرآزمونt2"]),
    ("t1_result",        ["نتیجهآزمونt1"]),
    ("t2_result",        ["نتیجهآزمونt2"]),
    ("avg_requirement",  ["آزمونالزاممیانگینبستهها"]),
    ("shortage",         ["درصدکمبودوزن", "درصدکمبودحجم"]),
    ("overall_result",   ["نتیجهکلیاستاندارد16381"]),
    ("declared_tolerance", ["رواداریاظهارشده"]),
]

# ──────────────────────────────────────────────
#  CSS سفارشی - اصلاح‌شده برای فونت فارسی
# ──────────────────────────────────────────────
CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Vazirmatn:wght@300;400;500;600;700;800;900&display=swap');

:root {
    --navy:     #0B1E3D;
    --navy2:    #112952;
    --blue:     #1A56DB;
    --sky:      #3B82F6;
    --teal:     #0EA5E9;
    --gold:     #F59E0B;
    --green:    #10B981;
    --red:      #EF4444;
    --orange:   #F97316;
    --bg:       #F0F4FF;
    --surface:  #FFFFFF;
    --border:   #CBD5E1;
    --text:     #1E293B;
    --muted:    #64748B;
    --radius:   14px;
    --shadow:   0 4px 24px rgba(11,30,61,0.10);
    --shadow-lg:0 8px 40px rgba(11,30,61,0.15);
}

/* ── فونت سراسری با اولویت بالا ── */
* {
    font-family: 'Vazirmatn', 'Tahoma', sans-serif !important;
}

html, body, [class*="css"], [class*="st-"], .stApp, .stMarkdown, 
.stMarkdown p, .stText, p, h1, h2, h3, h4, h5, h6, div, span, 
button, input, select, textarea, label, th, td, a, li, .stButton > button,
.stDownloadButton > button, [data-testid="stMetricValue"],
[data-testid="stMetricLabel"], .streamlit-expanderHeader,
.stTabs [data-baseweb="tab"], .stAlert, .stSuccess, .stWarning, 
.stError, .stInfo, .stSelectbox, [data-baseweb="select"],
[data-testid="stFileUploader"], .stSlider {
    font-family: 'Vazirmatn', 'Tahoma', sans-serif !important;
    direction: rtl !important;
}

/* ── صفحه پس‌زمینه ── */
.stApp {
    background: linear-gradient(135deg, #e8eef8 0%, #f0f4ff 60%, #e4ecf7 100%);
    min-height: 100vh;
}

/* ── هدر اصلی ── */
.app-header {
    background: linear-gradient(135deg, var(--navy) 0%, var(--navy2) 60%, #1a3a6e 100%);
    padding: 2.5rem 3rem 2rem;
    border-radius: 0 0 28px 28px;
    margin: -1rem -1rem 2rem -1rem;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow-lg);
}
.app-header::before {
    content: '';
    position: absolute;
    top: -80px; right: -80px;
    width: 320px; height: 320px;
    background: radial-gradient(circle, rgba(59,130,246,0.18) 0%, transparent 70%);
    border-radius: 50%;
}
.app-header::after {
    content: '';
    position: absolute;
    bottom: -60px; left: -60px;
    width: 240px; height: 240px;
    background: radial-gradient(circle, rgba(245,158,11,0.12) 0%, transparent 70%);
    border-radius: 50%;
}
.header-badge {
    display: inline-block;
    background: rgba(245,158,11,0.18);
    color: var(--gold);
    border: 1px solid rgba(245,158,11,0.35);
    padding: 4px 14px;
    border-radius: 20px;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 0.05em;
    margin-bottom: 0.8rem;
}
.header-title {
    color: #fff;
    font-size: 2rem;
    font-weight: 900;
    margin: 0 0 0.4rem;
    line-height: 1.2;
}
.header-sub {
    color: rgba(255,255,255,0.65);
    font-size: 0.92rem;
    font-weight: 400;
    margin: 0;
}

/* ── کارت‌های آمار ── */
.stat-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin: 1.5rem 0;
}
.stat-card {
    background: var(--surface);
    border-radius: var(--radius);
    padding: 1.4rem 1.2rem;
    box-shadow: var(--shadow);
    border-top: 4px solid var(--blue);
    position: relative;
    overflow: hidden;
    transition: transform 0.2s, box-shadow 0.2s;
}
.stat-card:hover { transform: translateY(-3px); box-shadow: var(--shadow-lg); }
.stat-card.green  { border-top-color: var(--green); }
.stat-card.orange { border-top-color: var(--orange); }
.stat-card.red    { border-top-color: var(--red); }
.stat-card.blue   { border-top-color: var(--blue); }
.stat-card .label {
    color: var(--muted); font-size: 0.78rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 0.5rem;
}
.stat-card .value {
    font-size: 2.2rem; font-weight: 900; line-height: 1;
    color: var(--text);
}
.stat-card .icon {
    position: absolute; left: 1.2rem; top: 50%;
    transform: translateY(-50%);
    font-size: 2rem; opacity: 0.12;
}

/* ── نشانک‌های وضعیت ── */
.badge {
    display: inline-flex; align-items: center; gap: 5px;
    padding: 4px 12px; border-radius: 20px;
    font-size: 0.78rem; font-weight: 700; white-space: nowrap;
}
.badge-green  { background:#D1FAE5; color:#065F46; }
.badge-orange { background:#FEF3C7; color:#92400E; }
.badge-red    { background:#FEE2E2; color:#991B1B; }
.badge-gray   { background:#F1F5F9; color:#475569; }
.badge-blue   { background:#DBEAFE; color:#1E40AF; }

/* ── جدول نتایج ── */
.results-table {
    background: var(--surface);
    border-radius: var(--radius);
    box-shadow: var(--shadow);
    overflow: hidden;
    margin: 1rem 0;
}
.results-table-header {
    background: var(--navy);
    padding: 1rem 1.5rem;
    display: flex; align-items: center; justify-content: space-between;
}
.results-table-header h3 { color: #fff; margin: 0; font-size: 1rem; font-weight: 700; }

/* ── کارت‌های جزئیات ── */
.detail-card {
    background: var(--surface);
    border-radius: var(--radius);
    padding: 1.5rem;
    box-shadow: var(--shadow);
    margin-bottom: 1rem;
    border-right: 4px solid var(--blue);
}
.detail-card h4 {
    margin: 0 0 1rem;
    font-size: 0.95rem; font-weight: 700; color: var(--navy);
    padding-bottom: 0.6rem;
    border-bottom: 1px solid var(--border);
}
.kv-row {
    display: flex; justify-content: space-between; align-items: flex-start;
    padding: 0.45rem 0; border-bottom: 1px solid #f1f5f9; font-size: 0.85rem;
}
.kv-row:last-child { border-bottom: none; }
.kv-key   { color: var(--muted); font-weight: 500; flex: 0 0 55%; }
.kv-val   { color: var(--text);  font-weight: 700; text-align: left; flex: 0 0 42%; }

/* ── نوار پیشرفت ── */
.progress-wrap {
    background: var(--surface); border-radius: var(--radius);
    padding: 1.5rem 2rem; box-shadow: var(--shadow); margin: 1rem 0;
}
.progress-label {
    display: flex; justify-content: space-between;
    font-size: 0.82rem; font-weight: 600; color: var(--muted);
    margin-bottom: 0.6rem;
}
.progress-bar-bg {
    height: 10px; background: #e2e8f0; border-radius: 6px; overflow: hidden;
}
.progress-bar-fill {
    height: 100%;
    background: linear-gradient(90deg, var(--sky), var(--blue));
    border-radius: 6px;
    transition: width 0.4s ease;
}

/* ── دکمه اصلی ── */
.stButton > button {
    background: linear-gradient(135deg, var(--blue) 0%, var(--navy) 100%) !important;
    color: white !important;
    border: none !important;
    padding: 0.75rem 2.5rem !important;
    border-radius: 10px !important;
    font-size: 1rem !important;
    font-weight: 700 !important;
    cursor: pointer !important;
    box-shadow: 0 4px 14px rgba(26,86,219,0.35) !important;
    transition: all 0.2s !important;
    width: 100% !important;
    letter-spacing: 0.02em !important;
}
.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 20px rgba(26,86,219,0.45) !important;
}

/* ── دانلود دکمه ── */
.stDownloadButton > button {
    background: linear-gradient(135deg, var(--green) 0%, #059669 100%) !important;
    color: white !important; border: none !important;
    padding: 0.65rem 2rem !important; border-radius: 10px !important;
    font-size: 0.92rem !important; font-weight: 700 !important;
    box-shadow: 0 4px 14px rgba(16,185,129,0.3) !important;
    width: 100% !important;
}

/* ── هشدار / موفقیت / خطا ── */
.stAlert, .stSuccess, .stWarning, .stError, .stInfo {
    border-radius: 10px !important;
}

/* ── تب‌ها ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px; background: transparent;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px !important; padding: 8px 20px !important;
    font-weight: 600 !important;
}

/* ── متریک ── */
[data-testid="metric-container"] {
    background: var(--surface);
    border-radius: var(--radius);
    padding: 1rem 1.2rem !important;
    box-shadow: var(--shadow);
}

/* ── خط تقسیم ── */
hr { border-color: var(--border) !important; margin: 1.5rem 0 !important; }

/* ── اسکرول‌بار ── */
::-webkit-scrollbar { width: 7px; height: 7px; }
::-webkit-scrollbar-track { background: #f1f5f9; border-radius: 4px; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }

/* ── فایل آپلودر ── */
[data-testid="stFileUploader"] {
    border-radius: var(--radius) !important;
    background: var(--surface) !important;
    box-shadow: var(--shadow) !important;
}

/* ── انیمیشن fade-in ── */
@keyframes fadeUp {
    from { opacity: 0; transform: translateY(18px); }
    to   { opacity: 1; transform: translateY(0); }
}
.fade-up { animation: fadeUp 0.45s ease both; }

/* ── نشانگر ریسک ── */
.risk-meter {
    position: relative; height: 8px;
    background: linear-gradient(90deg, #10B981, #F59E0B, #EF4444);
    border-radius: 4px; margin: 8px 0;
}
.risk-pointer {
    position: absolute; top: -4px;
    width: 16px; height: 16px;
    background: white; border: 3px solid var(--navy);
    border-radius: 50%; transform: translateX(-50%);
    transition: left 0.5s ease;
}

/* ── لوگو/عنوان sidebar ── */
.sidebar-logo {
    background: linear-gradient(135deg, var(--navy) 0%, var(--navy2) 100%);
    border-radius: var(--radius); padding: 1.2rem;
    text-align: center; margin-bottom: 1.5rem;
}
.sidebar-logo .logo-icon { font-size: 2.2rem; }
.sidebar-logo .logo-text { color: #fff; font-weight: 800; font-size: 0.95rem; margin-top: 0.4rem; }

/* ── اصلاح جهت متون Streamlit ── */
.stMarkdown, .stText, p, h1, h2, h3, h4, h5, h6 {
    text-align: right !important;
}

/* ── جهت RTL برای المنت‌های فرم ── */
[data-baseweb="select"] {
    direction: rtl !important;
}

/* ── دیتافریم Streamlit ── */
[data-testid="stDataFrame"] {
    font-family: 'Vazirmatn', 'Tahoma', sans-serif !important;
}
[data-testid="stDataFrame"] th {
    font-family: 'Vazirmatn', 'Tahoma', sans-serif !important;
    font-weight: 700 !important;
    text-align: right !important;
}
[data-testid="stDataFrame"] td {
    font-family: 'Vazirmatn', 'Tahoma', sans-serif !important;
    text-align: right !important;
}
</style>
"""

# ──────────────────────────────────────────────
#  توابع کمکی (بدون تغییر در منطق)
# ──────────────────────────────────────────────
def norm_text(s: Any) -> str:
    if s is None: return ""
    s = str(s).translate(PERSIAN_DIGITS)
    s = s.replace("ي","ی").replace("ك","ک").replace("\u200c","")
    s = re.sub(r"[\s\-\–\—\(\)\[\]\{\}\.,:;،/\\]+","",s)
    s = re.sub(r"[^\w\u0600-\u06FF]+","",s)
    return s.lower()

def to_float(v: Any) -> Optional[float]:
    if v is None or v == "": return None
    if isinstance(v, bool): return None
    if isinstance(v, (int,float,np.integer,np.floating)):
        return None if math.isnan(float(v)) else float(v)
    s = str(v).translate(PERSIAN_DIGITS).strip()
    s = s.replace(",",".")
    m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if not m: return None
    try: return float(m.group(0))
    except: return None

def fmt_num(x: Any, nd: int=4) -> str:
    if x is None: return "نامشخص"
    try: x=float(x)
    except: return str(x)
    s = f"{x:.{nd}f}".rstrip("0").rstrip(".")
    return s if s else "0"

def has_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")

def stdev_sample(xs: List[float]) -> Optional[float]:
    return statistics.stdev(xs) if len(xs) >= 2 else None

def find_header_row(ws, max_scan: int=20) -> int:
    best_score, best_row = -1, 1
    keywords = ["وزنناخالص","وزنبسته","وزنخالص","حجمخالص","خطایt1","خطایt2","نامفراورده"]
    for r in range(1, min(ws.max_row, max_scan)+1):
        joined = " ".join(norm_text(ws.cell(r,c).value) for c in range(1, min(ws.max_column,15)+1))
        score = sum(1 for kw in keywords if kw in joined)
        if score > best_score: best_score,best_row = score,r
    return best_row

def find_column_map(ws, header_row: int) -> Dict[str,int]:
    mapping: Dict[str,int] = {}
    for c in range(1, min(ws.max_column,20)+1):
        txt = norm_text(ws.cell(header_row,c).value)
        if not txt: continue
        for key,kws in COLUMN_KEYWORDS.items():
            if key in mapping: continue
            if any(kw in txt for kw in kws): mapping[key] = c
    return mapping

def find_sample_end(ws, header_row: int, col_map: Dict[str,int]) -> int:
    core_cols = [col_map[k] for k in ["gross","package","net_weight","net_volume","t1_error","t2_error"] if k in col_map]
    if not core_cols: core_cols = list(range(1,min(ws.max_column,8)+1))
    last = header_row
    for r in range(header_row+1, ws.max_row+1):
        if any(ws.cell(r,c).value not in (None,"") for c in core_cols): last = r
    return last

def extract_series(ws, wsv, col_idx: int, start_row: int, end_row: int):
    values: List[float] = []
    for r in range(start_row, end_row+1):
        num = to_float(wsv.cell(r,col_idx).value)
        if num is not None: values.append(num)
    return values, []

def pick_primary_series(series_meta: Dict[str,Dict[str,Any]]) -> Optional[str]:
    priority = {"gross":0,"package":1,"net_weight":2,"net_volume":3,"corrected_weight":4,"corrected_volume":5}
    best = None
    for key,meta in series_meta.items():
        xs = meta["values"]
        if len(xs) < 5: continue
        sdv = stdev_sample(xs)
        score = (len(xs), sdv if sdv is not None else -1, -priority.get(key,99))
        if best is None or score > best[0]: best = (score,key)
    return best[1] if best else None

def extract_nominal(sheet_name, summary_nominal=None, product_text=None, default_nominal=None):
    for source in [sheet_name, product_text or "", str(summary_nominal) if summary_nominal is not None else "", str(default_nominal) if default_nominal is not None else ""]:
        m = re.search(r"([0-9]+(?:\.[0-9]+)?)", str(source).translate(PERSIAN_DIGITS))
        if m: return float(m.group(0)), source
    return None, "نامشخص"

def get_correction_factor(n: int) -> float:
    if n >= 120: return 0.234
    elif n >= 80: return 0.295
    else: return 0.379

def runs_test(xs: List[float]):
    if len(xs) < 10: return None
    med = statistics.median(xs)
    signs = [1 if x>med else -1 if x<med else 0 for x in xs]
    signs = [s for s in signs if s != 0]
    n1 = sum(1 for s in signs if s==1)
    n2 = sum(1 for s in signs if s==-1)
    if n1==0 or n2==0: return {"runs":None,"z":None,"p":None}
    runs = 1+sum(1 for i in range(1,len(signs)) if signs[i]!=signs[i-1])
    mu  = 1+(2*n1*n2)/(n1+n2)
    var = (2*n1*n2*(2*n1*n2-n1-n2))/(((n1+n2)**2)*(n1+n2-1))
    z   = (runs-mu)/math.sqrt(var) if var>0 else None
    p   = 2*(1-stats.norm.cdf(abs(z))) if z is not None else None
    return {"runs":runs,"z":z,"p":p,"n1":n1,"n2":n2,"mu":mu}

def benford(xs: List[float]):
    digs = []
    for x in xs:
        a = abs(float(x))
        if a<=0: continue
        s = re.sub(r"[^0-9]","",f"{a:.15g}").lstrip("0")
        if s:
            d = int(s[0])
            if 1<=d<=9: digs.append(d)
    if len(digs)<20: return None
    obs = Counter(digs); n = len(digs)
    exp = {d: math.log10(1+1/d)*n for d in range(1,10)}
    chi2 = sum((obs.get(d,0)-exp[d])**2/exp[d] for d in range(1,10))
    return {"n":n,"obs":obs,"chi2":chi2,"p":float(stats.chi2.sf(chi2,df=8))}

def duplicate_stats(xs: List[float]):
    c = Counter(xs)
    dup = sum(v-1 for v in c.values() if v>1)
    return {"unique":len(c),"dup_count":dup,"dup_rate":dup/len(xs) if xs else None,"most_common":c.most_common(5)}

def decimal_stats(xs: List[float]):
    if not xs: return None
    dig1 = []
    for x in xs:
        s1 = f"{float(x):.1f}"
        dig1.append(s1.split(".",1)[1][-1])
    c1 = Counter(dig1)
    return {
        "n": len(xs),
        "pct_0": c1.get("0",0)/len(dig1)*100,
        "pct_5": c1.get("5",0)/len(dig1)*100,
        "pct_0_or_5": (c1.get("0",0)+c1.get("5",0))/len(dig1)*100,
    }

def lag1(xs: List[float]):
    if len(xs)<3: return None
    a,b = xs[:-1], xs[1:]
    if len(set(a))<2 or len(set(b))<2: return {"r":None,"p":None}
    r,p = stats.pearsonr(a,b)
    return {"r":float(r),"p":float(p)}

def iqr_outliers(xs: List[float]):
    if len(xs)<4: return None
    q1  = statistics.quantiles(xs,n=4,method="inclusive")[0]
    q3  = statistics.quantiles(xs,n=4,method="inclusive")[2]
    iqr = q3-q1
    lo,hi = q1-1.5*iqr, q3+1.5*iqr
    idx = [i for i,x in enumerate(xs) if x<lo or x>hi]
    return {"q1":q1,"q3":q3,"iqr":iqr,"low":lo,"high":hi,"count":len(idx),"indices":idx}

def shapiro_test(xs: List[float]):
    if len(xs)<3 or len(xs)>5000: return None
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        w,p = stats.shapiro(xs)
    return {"W":float(w),"p":float(p)}

def extreme_outliers(xs: List[float], n_sigma: float=3.0) -> int:
    if len(xs)<2: return 0
    mean = sum(xs)/len(xs); sd = stdev_sample(xs)
    if sd is None or sd==0: return 0
    return sum(1 for x in xs if abs(x-mean)>n_sigma*sd)

def find_summary(ws, wsv):
    found = {}
    for r in range(1, min(ws.max_row,40)+1):
        for c in range(1, min(ws.max_column,12)+1):
            label = ws.cell(r,c).value; keytxt = norm_text(label)
            if not keytxt: continue
            for key,kws in SUMMARY_PATTERNS:
                if key in found: continue
                if any(kw in keytxt for kw in kws):
                    vc = c-1 if c>1 else None
                    found[key] = {
                        "label": label, "label_cell": ws.cell(r,c).coordinate,
                        "value": wsv.cell(r,vc).value if vc else None,
                        "formula": ws.cell(r,vc).value if vc else None,
                    }
    return found

def detect_measure_type(series_meta, summary):
    if "net_volume" in series_meta and series_meta["net_volume"]["count"]>0: return "حجم"
    if "net_weight" in series_meta and series_meta["net_weight"]["count"]>0: return "وزن"
    return "وزن"

def expected_summary(series_meta, nominal, measure_type, sample_count):
    out={}
    package = series_meta.get("package",{}).get("values",[])
    net_key = "net_weight" if measure_type=="وزن" else "net_volume"
    net_vals = series_meta.get(net_key,{}).get("values",[])
    t1 = series_meta.get("t1_error",{}).get("values",[])
    t2 = series_meta.get("t2_error",{}).get("values",[])
    cf = get_correction_factor(sample_count)
    if package: out["avg_package"] = sum(package)/len(package)
    if net_vals:
        out["avg_net"] = sum(net_vals)/len(net_vals)
        if len(net_vals)>=2:
            out["sd_net"] = statistics.stdev(net_vals)
            out["corrected_avg"] = cf*out["sd_net"]+out["avg_net"]
    if t1:
        out["t1_fail_count"] = sum(1 for x in t1 if to_float(x) not in (None,0.0))
        out["t1_result"] = "قبول" if out["t1_fail_count"]<8 else "مردود"
    if t2:
        out["t2_fail_count"] = sum(1 for x in t2 if to_float(x) not in (None,0.0))
        out["t2_result"] = "قبول" if out["t2_fail_count"]<1 else "مردود"
    if "corrected_avg" in out and nominal is not None:
        out["avg_requirement"] = "قبول" if out["corrected_avg"]>=nominal else "مردود"
        if "avg_net" in out:
            out["shortage"] = 0 if out["avg_requirement"]=="قبول" else (nominal-out["avg_net"])/nominal*100
    if all(k in out for k in ["t1_result","t2_result","avg_requirement"]):
        out["overall_result"] = (out["t1_result"]=="قبول" and out["t2_result"]=="قبول" and out["avg_requirement"]=="قبول")
    return out

# ──────────────────────────────────────────────
#  پردازش فایل
# ──────────────────────────────────────────────
def process_file(path: Path, global_wb: Workbook, all_records: List[Dict]):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    wv = load_workbook(path, data_only=True)

    for sname in wb.sheetnames:
        ws = wb[sname]; dvs = wv[sname]
        header_row = find_header_row(ws)
        col_map    = find_column_map(ws, header_row)
        sample_end = find_sample_end(ws, header_row, col_map)
        sample_start = header_row+1
        sample_count = max(0, sample_end-sample_start+1)

        series_meta = {}
        for key in ["gross","package","net_weight","net_volume","corrected_weight","corrected_volume","t1_error","t2_error"]:
            if key in col_map:
                xs,_ = extract_series(ws,dvs,col_map[key],sample_start,sample_end)
                series_meta[key] = {"values":xs,"count":len(xs),"col":col_map[key]}

        primary = pick_primary_series(series_meta)
        primary_values = series_meta[primary]["values"] if primary else []

        sp = {"n":len(primary_values),"mean":None,"sd":None,"cv":None}
        if primary_values:
            sp["mean"] = sum(primary_values)/len(primary_values)
            sp["sd"]   = stdev_sample(primary_values)
            if sp["mean"] and sp["sd"] is not None: sp["cv"] = sp["sd"]/sp["mean"]*100

        run  = runs_test(primary_values)    if primary_values else None
        ben  = benford(primary_values)      if primary_values else None
        dup  = duplicate_stats(primary_values) if primary_values else None
        dec  = decimal_stats(primary_values)   if primary_values else None
        la   = lag1(primary_values)         if primary_values else None
        oiq  = iqr_outliers(primary_values) if primary_values else None
        shp  = shapiro_test(primary_values) if primary_values else None
        exc  = extreme_outliers(primary_values) if primary_values else 0

        summary     = find_summary(ws, dvs)
        product_txt = dvs["A2"].value if sample_start<=2<=sample_end else None
        nominal, nominal_src = extract_nominal(
            ws.title,
            summary.get("nominal",{}).get("value") if "nominal" in summary else None,
            product_txt, DEFAULT_NOMINAL)
        measure_type = detect_measure_type(series_meta, summary)
        exp = expected_summary(series_meta, nominal, measure_type, sample_count)

        reasons=[]; risk=0; critical=False

        if "gross" not in col_map or series_meta.get("gross",{}).get("count",0)==0:
            critical=True; reasons.append("ستون وزن ناخالص یافت نشد یا خالی است."); risk+=25

        net_key = "net_weight" if measure_type=="وزن" else "net_volume"
        if net_key not in col_map:
            critical=True; reasons.append(f"ستون {'وزن' if measure_type=='وزن' else 'حجم'} خالص پیدا نشد."); risk+=30

        for key,label in [("t1_error","خطای T1"),("t2_error","خطای T2")]:
            if key not in col_map: critical=True; reasons.append(f"ستون {label} پیدا نشد.")

        for key,label in [(net_key,"مقدار خالص"),("t1_error","خطای T1"),("t2_error","خطای T2")]:
            if key not in col_map: continue
            c = col_map[key]
            fr = [r for r in range(sample_start,sample_end+1) if to_float(dvs.cell(r,c).value) is not None and has_formula(ws.cell(r,c))]
            cov = len(fr)/sample_count if sample_count else 0
            if cov<0.95:
                reasons.append(f"پوشش فرمول ستون {label}: {len(fr)}/{sample_count} ردیف ({cov*100:.0f}%)")
                risk+=int((1-cov)*(35 if key==net_key else 25))

        def ck(key,calc,label,tol=1e-6):
            nonlocal risk
            if key in summary and calc is not None:
                rec = to_float(summary[key]["value"])
                if rec is not None and abs(rec-calc)>tol:
                    reasons.append(f"{label}: ثبت‌شده={fmt_num(rec)} | محاسبه={fmt_num(calc)}")
                    rel = abs(rec-calc)/max(abs(calc),1e-6)
                    risk+=20 if rel>0.05 else 10

        if "avg_package" in exp: ck("avg_package",exp["avg_package"],"میانگین وزن بسته")
        if "avg_net"     in exp: ck("avg_net",exp["avg_net"],f"میانگین {measure_type} خالص")
        if "sd_net"      in exp: ck("sd_net",exp["sd_net"],f"انحراف معیار {measure_type}")
        if "corrected_avg" in exp: ck("corrected_avg",exp["corrected_avg"],f"میانگین تصحیح‌شده")

        overall=None
        if "overall_result" in summary:
            ov = summary["overall_result"]["value"]
            overall = bool(ov) if isinstance(ov,bool) else str(ov).strip().lower() in ("true","1","قبول","yes")
        if overall is False: reasons.append("نتیجه کلی فرم: مردود"); risk+=10

        if exc>0: reasons.append(f"{exc} داده پرت شدید (بیش از ۳ انحراف معیار)"); risk+=min(30,exc*15)
        if nominal and "avg_net" in exp and exp["avg_net"] and abs(exp["avg_net"]-nominal)/nominal>0.10:
            reasons.append(f"میانگین {measure_type} خالص بیش از ۱۰٪ از مقدار اسمی فاصله دارد"); risk+=15

        if primary_values:
            if dup and dup["dup_rate"] and dup["dup_rate"]>0.3: reasons.append(f"نرخ تکرار بالا: {dup['dup_rate']*100:.1f}%"); risk+=8
            if run and run.get("p") and run["p"]<0.05:          reasons.append(f"Runs Test: p={run['p']:.4g} (غیرتصادفی)"); risk+=5
            if shp and shp.get("p") and shp["p"]<0.01:         reasons.append(f"توزیع غیرنرمال: W={shp['W']:.3f}, p={shp['p']:.4g}"); risk+=5
            if oiq and oiq["count"]>0:                          reasons.append(f"{oiq['count']} داده پرت IQR"); risk+=min(10,oiq["count"]*3)
            if dec and dec["pct_0_or_5"]>50:                    reasons.append(f"رقم دهم ۰ یا ۵: {dec['pct_0_or_5']:.1f}%"); risk+=5
            if ben and ben.get("p") and ben["p"]<0.001:         reasons.append("بنفورد: انحراف جدی از توزیع مورد انتظار"); risk+=2

        if primary is None:     critical=True; reasons.append("هیچ ستون عددی قابل‌اعتماد یافت نشد"); risk+=30
        if sample_count<MIN_ROWS_FOR_ANALYSIS: critical=True; reasons.append("تعداد نمونه کافی نیست"); risk+=20

        result = "خطا" if critical else ("بسیار مشکوک" if risk>=65 else "مشکوک" if risk>=35 else "سالم")

        all_records.append({
            "نام فایل": path.name, "نام شیت": sname,
            "تعداد نمونه": sample_count, "مقدار اسمی": nominal or "نامشخص",
            "نوع اندازه‌گیری": measure_type, "نتیجه نهایی": result,
            "امتیاز ریسک": min(100,int(round(risk))),
            "دلایل": " | ".join(reasons) if reasons else "مشکلی مشاهده نشد",
            "میانگین": fmt_num(sp["mean"]), "SD": fmt_num(sp["sd"]),
            "CV%": fmt_num(sp["cv"],2),
            "Runs p": fmt_num(run["p"],4) if run and run.get("p") else "—",
            "Benford p": fmt_num(ben["p"],4) if ben and ben.get("p") else "—",
            "Lag-1 r": fmt_num(la["r"],3) if la and la.get("r") else "—",
            "Shapiro p": fmt_num(shp["p"],4) if shp else "—",
            "IQR outlier": oiq["count"] if oiq else 0,
            "تکرار%": fmt_num(dup["dup_rate"]*100,1) if dup and dup.get("dup_rate") else "—",
            "رقم دهم ۰/۵%": fmt_num(dec["pct_0_or_5"],1) if dec else "—",
            "نتیجه فرم": "قبول" if overall is True else "مردود" if overall is False else "نامشخص",
        })

def build_excel_report(records: List[Dict], output_path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "گزارش خلاصه"
    headers = ["نام فایل","نام شیت","تعداد نمونه","مقدار اسمی","نوع اندازه‌گیری",
               "نتیجه نهایی","امتیاز ریسک","دلایل","میانگین","SD","CV%",
               "Runs p","Benford p","Lag-1 r","Shapiro p","IQR outlier","تکرار%","رقم دهم ۰/۵%","نتیجه فرم"]
    ws.append(headers)
    for rec in records:
        ws.append([rec.get(h) for h in headers])
    hf = PatternFill("solid",fgColor="0B1E3D"); hfnt = Font(color="FFFFFF",bold=True)
    for cell in ws[1]: cell.fill=hf; cell.font=hfnt; cell.alignment=Alignment(horizontal="center",wrap_text=True)
    gf=PatternFill("solid",fgColor="D1FAE5"); of=PatternFill("solid",fgColor="FEF3C7"); rf=PatternFill("solid",fgColor="FEE2E2")
    for row in ws.iter_rows(min_row=2):
        r=row[5].value
        fill = gf if r=="سالم" else of if r=="مشکوک" else rf if r in ("بسیار مشکوک","خطا") else None
        if fill:
            for cell in row: cell.fill=fill
    widths={1:28,2:18,3:12,4:12,5:14,6:14,7:10,8:60,9:14,10:12,11:10,12:12,13:12,14:12,15:12,16:12,17:10,18:12,19:14}
    for i,w in widths.items(): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; ws.sheet_view.rightToLeft=True
    wb.save(output_path)

# ──────────────────────────────────────────────
#  رابط کاربری
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="سامانه بازرسی ۱۶۳۸۱",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ── نوار کناری ──
with st.sidebar:
    st.markdown("""
    <div class="sidebar-logo">
        <div class="logo-icon">🔬</div>
        <div class="logo-text">سامانه بازرسی ISIRI 16381</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("#### ⚙️ تنظیمات")
    min_rows = st.slider("حداقل تعداد نمونه", 5, 50, MIN_ROWS_FOR_ANALYSIS, step=5)
    risk_threshold_high = st.slider("آستانه ریسک بالا", 40, 90, 65, step=5)
    risk_threshold_med  = st.slider("آستانه ریسک متوسط", 10, 60, 35, step=5)

    st.markdown("---")
    st.markdown("#### 📌 راهنمای نتایج")
    st.markdown("""
    <div style="font-size:0.82rem; line-height:2">
    <span class="badge badge-green">✔ سالم</span> ریسک پایین<br>
    <span class="badge badge-orange">⚠ مشکوک</span> نیاز به بررسی<br>
    <span class="badge badge-red">✖ بسیار مشکوک</span> پرریسک<br>
    <span class="badge badge-gray">⊘ خطا</span> ساختار ناشناخته
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    st.caption("نسخه ۲.۱ | استاندارد ۱۶۳۸۱")

# ── هدر ──
st.markdown("""
<div class="app-header fade-up">
    <div class="header-badge">ISIRI 16381 · بازرسی داده</div>
    <h1 class="header-title">سامانه بازرسی پیشرفته فرم‌های کنترل پیش‌بسته‌بندی</h1>
    <p class="header-sub">تشخیص داده‌سازی مصنوعی · اعتبارسنجی فرمول‌ها · تحلیل آماری audit-grade</p>
</div>
""", unsafe_allow_html=True)

# ── ناحیه آپلود ──
st.markdown('<div class="fade-up">', unsafe_allow_html=True)

col_up, col_info = st.columns([3, 1])
with col_up:
    uploaded_files = st.file_uploader(
        "فایل‌های اکسل را اینجا رها کنید",
        type="xlsx",
        accept_multiple_files=True,
        label_visibility="visible",
        help="فرمت xlsx · چندین فایل همزمان قابل انتخاب است"
    )

with col_info:
    st.markdown("""
    <div style="background:#EFF6FF;border-radius:12px;padding:1.2rem;border:1px solid #BFDBFE;margin-top:0.5rem">
        <div style="font-size:0.85rem;font-weight:700;color:#1E40AF;margin-bottom:0.6rem">📋 نکات مهم</div>
        <div style="font-size:0.78rem;color:#1E40AF;line-height:1.8">
        • فایل خام و پرشده هر دو قابل پردازش‌اند<br>
        • چند شیت همزمان بررسی می‌شود<br>
        • گزارش Excel قابل دانلود است<br>
        • داده‌ها پیش از آپلود رمزگذاری نمی‌شوند
        </div>
    </div>
    """, unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)

# ── دکمه پردازش ──
run_btn = st.button("🚀  شروع پردازش و تحلیل", use_container_width=True)

if run_btn:
    if not uploaded_files:
        st.error("⚠️ لطفاً حداقل یک فایل Excel انتخاب کنید.")
        st.stop()

    # ── نوار پیشرفت زیبا ──
    progress_container = st.empty()
    total = len(uploaded_files)

    global_wb  = Workbook(); global_wb.remove(global_wb.active)
    all_records: List[Dict] = []

    for idx, uf in enumerate(uploaded_files):
        pct = int((idx / total) * 100)
        progress_container.markdown(f"""
        <div class="progress-wrap fade-up">
            <div class="progress-label">
                <span>📂 در حال پردازش: <b>{uf.name}</b></span>
                <span>{idx+1} از {total} فایل · {pct}%</span>
            </div>
            <div class="progress-bar-bg">
                <div class="progress-bar-fill" style="width:{pct}%"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uf.getvalue())
            tmp_path = Path(tmp.name)
        try:
            process_file(tmp_path, global_wb, all_records)
        except Exception as e:
            all_records.append({
                "نام فایل": uf.name, "نام شیت": "خطا",
                "تعداد نمونه": 0, "مقدار اسمی": "—",
                "نوع اندازه‌گیری": "—", "نتیجه نهایی": "خطا",
                "امتیاز ریسک": 100,
                "دلایل": f"خطا در پردازش: {e}",
                "میانگین":"—","SD":"—","CV%":"—","Runs p":"—","Benford p":"—",
                "Lag-1 r":"—","Shapiro p":"—","IQR outlier":0,"تکرار%":"—","رقم دهم ۰/۵%":"—","نتیجه فرم":"—",
            })
        finally:
            tmp_path.unlink(missing_ok=True)

    progress_container.markdown(f"""
    <div class="progress-wrap fade-up">
        <div class="progress-label">
            <span>✅ پردازش کامل شد</span><span>100%</span>
        </div>
        <div class="progress-bar-bg">
            <div class="progress-bar-fill" style="width:100%;background:linear-gradient(90deg,#10B981,#059669)"></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── کارت‌های آمار خلاصه ──
    n_total   = len(all_records)
    n_safe    = sum(1 for r in all_records if r["نتیجه نهایی"]=="سالم")
    n_susp    = sum(1 for r in all_records if r["نتیجه نهایی"]=="مشکوک")
    n_high    = sum(1 for r in all_records if r["نتیجه نهایی"] in ("بسیار مشکوک","خطا"))
    avg_risk  = int(sum(r["امتیاز ریسک"] for r in all_records) / max(n_total,1))

    st.markdown(f"""
    <div class="stat-grid fade-up">
        <div class="stat-card blue">
            <div class="label">کل شیت‌های پردازش‌شده</div>
            <div class="value">{n_total}</div>
            <div class="icon">📄</div>
        </div>
        <div class="stat-card green">
            <div class="label">سالم</div>
            <div class="value">{n_safe}</div>
            <div class="icon">✔</div>
        </div>
        <div class="stat-card orange">
            <div class="label">مشکوک</div>
            <div class="value">{n_susp}</div>
            <div class="icon">⚠</div>
        </div>
        <div class="stat-card red">
            <div class="label">بسیار مشکوک / خطا</div>
            <div class="value">{n_high}</div>
            <div class="icon">✖</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── تب‌ها ──
    tab1, tab2, tab3 = st.tabs(["📋  جدول کامل نتایج", "🔍  جزئیات هر شیت", "📥  دانلود گزارش"])

    # ── تب ۱: جدول کامل با st.dataframe (اصلاح‌شده) ──
    with tab1:
        st.markdown("""
        <div class="results-table fade-up">
            <div class="results-table-header">
                <h3>📋 نتایج بازرسی</h3>
                <span style="color:rgba(255,255,255,0.6);font-size:0.8rem">{} شیت · {} فایل</span>
            </div>
        </div>
        """.format(n_total, len(uploaded_files)), unsafe_allow_html=True)

        # ساخت DataFrame برای نمایش
        df_display = pd.DataFrame(all_records)
        
        # انتخاب و مرتب‌سازی ستون‌ها برای نمایش
        display_columns = [
            "نام فایل", "نام شیت", "تعداد نمونه", "مقدار اسمی", 
            "نوع اندازه‌گیری", "نتیجه نهایی", "امتیاز ریسک",
            "میانگین", "SD", "CV%", "Runs p", "Benford p",
            "Lag-1 r", "Shapiro p", "IQR outlier", "تکرار%",
            "رقم دهم ۰/۵%", "نتیجه فرم", "دلایل"
        ]
        
        # فیلتر ستون‌های موجود
        available_columns = [col for col in display_columns if col in df_display.columns]
        df_show = df_display[available_columns].copy()

        # اعمال استایل شرطی با تابع کمکی
        def highlight_result(val):
            if val == "سالم":
                return 'background-color: #D1FAE5; color: #065F46; font-weight: bold'
            elif val == "مشکوک":
                return 'background-color: #FEF3C7; color: #92400E; font-weight: bold'
            elif val in ("بسیار مشکوک", "خطا"):
                return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold'
            return ''

        def highlight_risk(val):
            try:
                v = int(val)
                if v < 35:
                    return 'background-color: #D1FAE5; color: #065F46'
                elif v < 65:
                    return 'background-color: #FEF3C7; color: #92400E'
                else:
                    return 'background-color: #FEE2E2; color: #991B1B'
            except:
                return ''

        # اعمال استایل
        styled_df = df_show.style.applymap(
            highlight_result, 
            subset=['نتیجه نهایی']
        ).applymap(
            highlight_risk, 
            subset=['امتیاز ریسک']
        ).set_properties(**{
            'text-align': 'right',
            'font-family': 'Vazirmatn, Tahoma, sans-serif'
        })

        # نمایش جدول
        st.dataframe(
            styled_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "دلایل": st.column_config.TextColumn(
                    "دلایل",
                    width="large",
                ),
                "امتیاز ریسک": st.column_config.ProgressColumn(
                    "امتیاز ریسک",
                    format="%d",
                    min_value=0,
                    max_value=100,
                ),
            }
        )

    # ── تب ۲: جزئیات (اصلاح‌شده برای نمایش کامل داده‌ها) ──
    with tab2:
        if not all_records:
            st.info("نتیجه‌ای برای نمایش وجود ندارد.")
        else:
            sheet_options = [f"{r['نام فایل']} › {r['نام شیت']}" for r in all_records]
            selected = st.selectbox("یک شیت را انتخاب کنید:", sheet_options)
            idx_sel  = sheet_options.index(selected)
            rec      = all_records[idx_sel]

            verdict = rec["نتیجه نهایی"]
            badge_map = {
                "سالم":         ("badge-green",  "✔ سالم",         "#10B981"),
                "مشکوک":        ("badge-orange", "⚠ مشکوک",        "#F97316"),
                "بسیار مشکوک": ("badge-red",    "✖ بسیار مشکوک",  "#EF4444"),
                "خطا":          ("badge-gray",   "⊘ خطا",           "#94A3B8"),
            }
            b_cls, b_txt, b_clr = badge_map.get(verdict, ("badge-gray","نامشخص","#94A3B8"))
            risk_val = rec["امتیاز ریسک"]

            # ── header کارت ──
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,#0B1E3D,#112952);border-radius:16px;padding:1.6rem 2rem;margin-bottom:1.2rem" class="fade-up">
                <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:1rem">
                    <div>
                        <div style="color:rgba(255,255,255,0.6);font-size:0.78rem;margin-bottom:4px">{rec['نام فایل']}</div>
                        <div style="color:#fff;font-size:1.3rem;font-weight:800">{rec['نام شیت']}</div>
                    </div>
                    <div style="display:flex;align-items:center;gap:1.5rem">
                        <div style="text-align:center">
                            <div style="color:rgba(255,255,255,0.55);font-size:0.72rem;margin-bottom:2px">امتیاز ریسک</div>
                            <div style="color:{b_clr};font-size:2rem;font-weight:900">{risk_val}</div>
                        </div>
                        <div>
                            <span class="badge {b_cls}" style="font-size:0.9rem;padding:7px 18px">{b_txt}</span>
                        </div>
                    </div>
                </div>
                <div style="margin-top:1.2rem">
                    <div class="risk-meter">
                        <div class="risk-pointer" style="left:{min(risk_val,100)}%"></div>
                    </div>
                    <div style="display:flex;justify-content:space-between;color:rgba(255,255,255,0.4);font-size:0.7rem;margin-top:4px">
                        <span>۰ - سالم</span><span>۳۵ - مشکوک</span><span>۶۵ - پرریسک</span><span>۱۰۰</span>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            col_a, col_b, col_c = st.columns(3)

            with col_a:
                st.markdown(f"""
                <div class="detail-card fade-up">
                    <h4>📊 آمار پایه</h4>
                    <div class="kv-row"><span class="kv-key">تعداد نمونه</span><span class="kv-val">{rec.get('تعداد نمونه', '—')}</span></div>
                    <div class="kv-row"><span class="kv-key">مقدار اسمی</span><span class="kv-val">{rec.get('مقدار اسمی', '—')}</span></div>
                    <div class="kv-row"><span class="kv-key">نوع اندازه‌گیری</span><span class="kv-val">{rec.get('نوع اندازه‌گیری', '—')}</span></div>
                    <div class="kv-row"><span class="kv-key">میانگین محاسبه‌شده</span><span class="kv-val">{rec.get('میانگین', '—')}</span></div>
                    <div class="kv-row"><span class="kv-key">انحراف معیار (SD)</span><span class="kv-val">{rec.get('SD', '—')}</span></div>
                    <div class="kv-row"><span class="kv-key">ضریب تغییرات (CV%)</span><span class="kv-val">{rec.get('CV%', '—')}</span></div>
                    <div class="kv-row"><span class="kv-key">نتیجه کلی فرم</span><span class="kv-val">{rec.get('نتیجه فرم', '—')}</span></div>
                </div>
                """, unsafe_allow_html=True)

            with col_b:
                def pval_badge(p_str):
                    try:
                        p = float(p_str)
                        if p < 0.01:  return f'<span class="badge badge-red">{p_str} ⚠</span>'
                        if p < 0.05:  return f'<span class="badge badge-orange">{p_str} !</span>'
                        return f'<span class="badge badge-green">{p_str} ✔</span>'
                    except:
                        return f'<span class="badge badge-gray">{p_str}</span>'

                st.markdown(f"""
                <div class="detail-card fade-up">
                    <h4>🧪 آزمون‌های آماری</h4>
                    <div class="kv-row"><span class="kv-key">Runs Test p-value</span><span class="kv-val">{pval_badge(rec.get('Runs p', '—'))}</span></div>
                    <div class="kv-row"><span class="kv-key">Benford p-value</span><span class="kv-val">{pval_badge(rec.get('Benford p', '—'))}</span></div>
                    <div class="kv-row"><span class="kv-key">Lag-1 autocorr (r)</span><span class="kv-val">{rec.get('Lag-1 r', '—')}</span></div>
                    <div class="kv-row"><span class="kv-key">Shapiro-Wilk p</span><span class="kv-val">{pval_badge(rec.get('Shapiro p', '—'))}</span></div>
                    <div class="kv-row"><span class="kv-key">IQR Outliers</span><span class="kv-val">{rec.get('IQR outlier', '—')}</span></div>
                    <div class="kv-row"><span class="kv-key">نرخ تکرار</span><span class="kv-val">{rec.get('تکرار%', '—')}%</span></div>
                    <div class="kv-row"><span class="kv-key">رقم دهم ۰ یا ۵</span><span class="kv-val">{rec.get('رقم دهم ۰/۵%', '—')}%</span></div>
                </div>
                """, unsafe_allow_html=True)

            with col_c:
                reasons_html = ""
                if rec.get("دلایل") and rec["دلایل"] != "مشکلی مشاهده نشد":
                    for i, reason in enumerate(rec["دلایل"].split(" | "), 1):
                        reasons_html += f'<div style="padding:0.5rem 0;border-bottom:1px solid #f1f5f9;font-size:0.8rem;color:#1E293B"><span style="color:#EF4444;font-weight:700;margin-left:6px">#{i}</span>{reason}</div>'
                else:
                    reasons_html = '<div style="color:#10B981;font-weight:600;padding:1rem 0">✔ هیچ مشکلی شناسایی نشد</div>'

                st.markdown(f"""
                <div class="detail-card fade-up" style="border-right-color:#EF4444">
                    <h4>⚠️ دلایل هشدار</h4>
                    {reasons_html}
                </div>
                """, unsafe_allow_html=True)

            # ── نمودارها (اگر داده خام در رکورد ذخیره شده باشد) ──
            # توجه: برای سادگی، بخش نمودارها را غیرفعال می‌کنیم چون داده‌های خام
            # در all_records ذخیره نمی‌شوند. در صورت نیاز می‌توان بعداً اضافه کرد.
            st.info("💡 برای مشاهده نمودارهای توزیع و توالی، لطفاً فایل را مجدداً با قابلیت ذخیره‌سازی داده‌های خام پردازش کنید.")

    # ── تب ۳: دانلود ──
    with tab3:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#F0FDF4,#DCFCE7);border:1.5px solid #BBF7D0;
                    border-radius:16px;padding:2rem;text-align:center;margin-bottom:1.5rem" class="fade-up">
            <div style="font-size:2.5rem;margin-bottom:0.8rem">📥</div>
            <div style="font-size:1.2rem;font-weight:800;color:#065F46;margin-bottom:0.5rem">دانلود گزارش نهایی</div>
            <div style="color:#047857;font-size:0.88rem">
                فایل Excel کامل با فرمت‌بندی رنگی، فیلتر خودکار و راهنمای تفسیر
            </div>
        </div>
        """, unsafe_allow_html=True)

        c1, c2, c3 = st.columns([1,2,1])
        with c2:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
                build_excel_report(all_records, Path(tmp_out.name))
                with open(tmp_out.name, "rb") as f:
                    excel_bytes = f.read()
            Path(tmp_out.name).unlink(missing_ok=True)

            st.download_button(
                label="📥  دانلود گزارش Excel",
                data=excel_bytes,
                file_name="گزارش_بازرسی_16381.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # ── خلاصه در تب دانلود ──
        st.markdown("---")
        st.markdown("**📊 خلاصه آنچه در گزارش است:**")
        summary_items = [
            ("📋", "جدول کامل نتایج", f"{n_total} سطر با رنگ‌بندی خودکار"),
            ("✔", "امتیاز ریسک",      "۰ تا ۱۰۰ با کدگذاری رنگی"),
            ("🧪", "آزمون‌های آماری", "Runs، Benford، Shapiro، IQR، Lag-1"),
            ("⚠️", "دلایل هشدار",      "متن کامل هر هشدار با مقادیر عددی"),
            ("📐", "تطبیق فرم",         "مقایسه محاسبه مستقل با ثبت‌شده در فرم"),
        ]
        cols = st.columns(len(summary_items))
        for col, (icon, title, desc) in zip(cols, summary_items):
            with col:
                st.markdown(f"""
                <div style="background:#F8FAFC;border-radius:10px;padding:1rem;text-align:center;border:1px solid #E2E8F0">
                    <div style="font-size:1.6rem">{icon}</div>
                    <div style="font-weight:700;font-size:0.82rem;color:#1E293B;margin:4px 0">{title}</div>
                    <div style="font-size:0.74rem;color:#64748B">{desc}</div>
                </div>
                """, unsafe_allow_html=True)